from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import Color, PatternFill, Font, Border
redFill = PatternFill(start_color='EE1111',
               end_color='EE1111',
              fill_type='solid')

wb=load_workbook(r'C:\Users\india\Downloads\Sample - Superstore Sales (Excel) (1).xlsx')
ws=wb.get_sheet_by_name('Orders1')
wr=wb.get_sheet_by_name('Returns')
ns=wb.create_sheet('New Orders')
no=ws.max_row
print (no)
cname=[]
oname=[]
v=[]
noc=ws.max_column
for i in range(1,noc+1):
    if ws.cell(row=1,column=i).value=="Customer Name":
        cnameref=i
    elif ws.cell(row=1,column=i).value=="Order ID":
        nf=i
    elif ws.cell(row=1,column=i).value=="Order Date":
        df=i
for i in range(no):
        cname.append(ws.cell(row=i+2,column=cnameref).value)
        oname.append(ws.cell(row=i+2,column=nf).value)
        v.append(ws.cell(row=i+2, column=df).value)
print (cname,oname,v)
ws.conditional_formatting.add('B',CellIsRule(operator='lessThan', formula=['B1>250'], stopIfTrue=True, fill=redFill))


# ns['A1']= "Customer Name"
# ns['B1'] = "Order ID"
# ns['C1'] = "Order Date"
# ns['D1'] = "Returned Orders"
# for i in range(no):
#         ns.cell(row=i+2,column=1).value=cname[i]
#         ns.cell(row=i + 2, column=2).value = oname[i]
#         ns.cell(row=i + 2, column=3).value= v[i]
#
# for i in range(ns.max_row):
#     for j in range(wr.max_row):
#         print (ns["B"+str(i+2)].value)
#         print(wr['A'+str(j+2)].value)
#         if ns["B"+str(i+2)].value==wr['A'+str(j+2)].value:
#             ns["D" + str(i + 2)]=wr["B" + str(j+2)].value

# for row in ns['B1:B15']:
#     for cell in row:
#        cell.value = "=VLOOKUP(D{0}, wr['A1:B15'], 1, 0)".format(cell.row)
# print (wb.sheetnames)
# list=[ 'Customer Name', 'Order Id','Order Date', 'Order Quantity', 'Unit price', 'Sales',' Profit',' Product Category']
# ws=wb.create_sheet('Orders')
# a=[i.value for c in ws['A1':'U1']for i in c if i.value in list]
# for i in range(0,len(list)):
#     wu[get_column_letter(i+1)+str(1)]=list[i]
#     print (wu[get_column_letter(i+1)+str(1)].value)
#
# for i in range (1,21):
#     for j in range(1,wu.max_column):
#         while ws.cell(row=1,column=i).value==wu.cell(row=1,column=j).value:
#             for k in range(ws.max_row):
#                 wu['A'+str(k+2)]=ws[get_column_letter(i)+str(k+2)].value
#
# ##for i i  range(len(ncol
# ##t=s.cell['A1'].value
# ##print (t)
#
# ##    wb.active()
# ##    wb.createsheet(title=h
# ##print(s.cell(row=0,column=10).value)
# ##for each in k:
# ##    print (each.value)
# ##s=sheet["Orders"]
# ##for row in s:
# ##    for cell in row:
# ##           wu[cell.coordinate].value=cell.value
wb.save("new.xlsx")
#
